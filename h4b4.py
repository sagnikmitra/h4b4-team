import streamlit as st
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# Load or create the workbook and sheet
def load_excel(file_name='participants.xlsx'):
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = ["Timestamp", "Name", "Email", "Phone", "CreateOrJoin", "TeamName", "GitHub", "LinkedIn"]
        sheet.append(headers)
        wb.save(file_name)
    return wb, sheet

# Get existing data
def get_existing_data(sheet):
    data = {
        "emails": set(),
        "phones": set(),
        "teams": set()
    }
    for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
        if row[2]:  # Email
            data["emails"].add(row[2].strip().lower())
        if row[3]:  # Phone
            data["phones"].add(row[3].strip())
        if row[5]:  # TeamName
            data["teams"].add(row[5].strip().lower())
    return data

# Check team strength
def get_team_members_count(team_name, sheet):
    team_members = [row[5] for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True) if row[5]]
    return team_members.count(team_name.lower())

# Validate and save the data to the Excel file
def validate_and_save(name, email, phone, action, team_name, github, linkedin, sheet, wb, existing_data):
    # Check for unique email and phone
    if email.lower() in existing_data["emails"]:
        st.error(f"Error: The email '{email}' is already registered.")
        return False
    if phone in existing_data["phones"]:
        st.error(f"Error: The phone number '{phone}' is already registered.")
        return False

    # Get all existing team names
    teams = existing_data["teams"]

    # Validation process
    if action == "CreateTeam":
        if len(team_name) < 8:
            st.error("Error: Team name must be at least 8 characters long.")
            return False
        if "team" in team_name.lower():
            st.error("Error: Team name cannot contain the word 'team'.")
            return False
        if any(team_name.lower() in t.lower() or t.lower() in team_name.lower() for t in teams):
            st.error("Error: Team name already exists or is too similar to another team name.")
            return False
        else:
            st.session_state.success_message = f"Hi {name}, you are registered successfully for Hack4Bengal Season 4 for Team {team_name}."
            st.session_state.is_registered = True
    elif action == "JoinTeam":
        # Check if the team exists
        if not any(team_name.lower() == t for t in teams):
            st.error(f"Error: No team named '{team_name}' found.")
            return False
        else:
            # Check team strength (max 4 members)
            team_count = get_team_members_count(team_name, sheet)
            if team_count >= 4:
                st.error(f"Error: Team '{team_name}' already has 4 members. You cannot join.")
                return False
            else:
                st.session_state.success_message = f"Hi {name}, you are registered successfully for Hack4Bengal Season 4 for Team {team_name}."
                st.session_state.is_registered = True
    else:
        st.error(f"Error: Invalid action '{action}'. Please select either 'CreateTeam' or 'JoinTeam'.")
        return False

    # If validation passes, append the data to the Excel file
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_row = [timestamp, name, email, phone, action, team_name, github, linkedin]
        sheet.append(new_row)
        wb.save('participants.xlsx')
        st.info("Data saved successfully!")
        return True
    except Exception as e:
        st.error(f"An error occurred while saving the data: {e}")
        return False

# Initialize session state
if 'name' not in st.session_state:
    st.session_state.name = ""
if 'email' not in st.session_state:
    st.session_state.email = ""
if 'phone' not in st.session_state:
    st.session_state.phone = ""
if 'action' not in st.session_state:
    st.session_state.action = "CreateTeam"  # Default value
if 'team_name' not in st.session_state:
    st.session_state.team_name = ""
if 'github' not in st.session_state:
    st.session_state.github = ""
if 'linkedin' not in st.session_state:
    st.session_state.linkedin = ""
if 'success_message' not in st.session_state:
    st.session_state.success_message = ""
if 'is_registered' not in st.session_state:
    st.session_state.is_registered = False

# Load the Excel sheet
wb, sheet = load_excel()
existing_data = get_existing_data(sheet)

# Streamlit UI
st.header("Hack4Bengal Season 4 Registration")
st.write("Please fill in your details to create or join a team.")

# Display success message if registration is successful
if st.session_state.is_registered:
    st.success(st.session_state.success_message)
    if st.button("Enter New Data"):
        # Reset session state for a new form entry
        st.session_state.name = ""
        st.session_state.email = ""
        st.session_state.phone = ""
        st.session_state.action = "CreateTeam"  # Reset to default value
        st.session_state.team_name = ""
        st.session_state.github = ""
        st.session_state.linkedin = ""
        st.session_state.is_registered = False
        st.experimental_rerun()
else:
    # Input fields
    st.session_state.name = st.text_input("Name", value=st.session_state.name)
    st.session_state.email = st.text_input("Email", value=st.session_state.email)
    st.session_state.phone = st.text_input("Phone", value=st.session_state.phone)
    st.session_state.action = st.radio(
        "Would you like to create or join a team?",
        ["CreateTeam", "JoinTeam"],
        index=["CreateTeam", "JoinTeam"].index(st.session_state.action) if st.session_state.action else 0
    )
    st.session_state.team_name = st.text_input("Team Name (Case Sensitive)", value=st.session_state.team_name)
    st.session_state.github = st.text_input("GitHub Profile (optional)", value=st.session_state.github)
    st.session_state.linkedin = st.text_input("LinkedIn Profile (optional)", value=st.session_state.linkedin)

    # Submit button
    if st.button("Submit"):
        # Ensure that mandatory fields are filled
        if not st.session_state.name or not st.session_state.email or not st.session_state.phone or not st.session_state.action or not st.session_state.team_name:
            st.error("Please fill in all the mandatory fields.")
        else:
            # Validate and save the data
            if validate_and_save(
                st.session_state.name,
                st.session_state.email,
                st.session_state.phone,
                st.session_state.action,
                st.session_state.team_name,
                st.session_state.github,
                st.session_state.linkedin,
                sheet,
                wb,
                existing_data
            ):
                st.session_state.success_message = f"Hi {st.session_state.name}, you are registered successfully for Hack4Bengal Season 4 for Team {st.session_state.team_name}. Click below to enter new data."
                st.session_state.is_registered = True
                st.experimental_rerun()
